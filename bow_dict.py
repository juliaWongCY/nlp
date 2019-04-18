import openpyxl
import dictionary as dic
import sys

SETS = ["Set1", "Set2", "Set3", "Set4", "Set5"]

if __name__ == "__main__":
    if (len(sys.argv) != 2):
        print "excel file required"
        sys.exit()
    
    wb = openpyxl.load_workbook(sys.argv[1])
    d = dic.Dictionary()

    fp = open("mpqa-subj-lexicon.tff")
    for i, line in enumerate(fp):
        l = line.split()
        word = l[2][6:]
        ispos = l[-1][-8:]
        if ispos == "=neutral":
            continue
        d.add_word((word, ispos == "positive"))


    for SET in SETS:
        sheet = wb[SET]
        count = 0
        for i in range(2, sheet.max_row):
            realscore = int(sheet['N' + str(i)].value * 4)
            comment = unicode(sheet['I' + str(i)].value).lower()
            score = 0
            #print
            for word in comment.split():
                score += d.find_word(word)
            #    print word + " " + str(d.find_word(word))
            if (score > 0) != (realscore < 0) or realscore == score == 0:
                count += 1
            #else:
            #    print
            #    print comment
            #    print realscore
            #    print score
        print SET + ": " + str(count / float(sheet.max_row - 1))

