import sys
import re
import pandas as pd
import openpyxl
from sets import Set

CONNECTIVES = ["not", "very", "extremely", "on"]
MIN_OCCURRENCE = 20
SETS = ["Set1", "Set2", "Set3", "Set4", "Set5"]

# converts string into list of words + count
def bag_of_words(s):
    s = s.lower()
    split_string = sorted(ngram_split(re.split("[^a-zA-Z0-9\']", s)))
    split_string = list(filter(lambda a: a != "", split_string))
    return split_string

# splits words based on connectives instead of just spaces
def ngram_split(arr):
    i = 0
    while i < len(arr):
        ngram_join(arr, i)
        i = i + 1
    return arr

# helper func
def ngram_join(arr, index):
    if (arr[index] in CONNECTIVES or arr[index].endswith("n't")) \
     and index < len(arr) - 1:
        ngram_join(arr, index + 1)
        word = arr.pop(index)
        arr[index] = word + " " + arr[index]

# counts the word occurrences in sheet, and the overall sentiment that 
# occurrence is associated with (-1, 0 or +1)
# word_set: the set of words already in dataframe df
def bag_train(df, sheet, word_set):
    for i in range(2, sheet.max_row + 1):
        comment = unicode(sheet['i' + str(i)].value)
        score = int(sheet['n' + str(i)].value * 4)
        score = min(1, score) if score > 0 else max(-1, score)
        bow = bag_of_words(comment)
        for word in bow:
            if word not in word_set:
                word_set.add(word)
                df.loc[word] = [0] * 3
                df.ix[word][score] += 1
            else:
                df.ix[word][score] += 1

# removes all rows in dataframe df with a sum less than MIN_OCCURRENCE
def bag_filter(df):
    df.drop(df[df.sum(axis=1) < MIN_OCCURRENCE].index, inplace=True)

# calculates percentage composition of each row
def bag_composition(df):
    for i in range(df.shape[0]):
        row_sum = 0
        for j in list(df):
            row_sum += df[j][i]
        for j in list(df):
            df[j][i] = round(df[j][i] / float(row_sum), 2)

# tests sheet with list of good and bad words
def bag_test(goods, bads, sheet):
    count = 0
    for i in range(2, sheet.max_row + 1):
        score = 0
        comment = unicode(sheet['i' + str(i)].value)
        realscore = int(sheet['n' + str(i)].value * 4)
        realscore = min(1, realscore) if score > 0 else max(-1, realscore)
        bow = bag_of_words(comment)
        for word in bow:
            if word in goods:
                score += 1
            elif word in bads:
                score -= 1
        if ((score > 0) != (realscore < 0)) or (score == realscore == 0):
            count += 1
        #else:
        #    print
        #    print comment
        #    print realscore
        #    print score
    return float(count) / (sheet.max_row - 1)

# single comment test
def bag_test_comment(goods, bads, comment):
    score = 0
    bow = bag_of_words(comment)
    for word in bow:
        if word in goods:
            score += 1
        elif word in bads:
            score -= 1
    print score

def df_print_full(df):
    pd.set_option("display.max_rows", len(df))
    print df
    pd.reset_option("display.max_rows")


if __name__ == "__main__":
    if (len(sys.argv) != 2):
        print "excel file required"
        sys.exit()
    
    wb = openpyxl.load_workbook(sys.argv[1])
    
    for j in range(5):
        df = pd.DataFrame(columns=range(-1, 2))
        word_set = Set()
        for i in [x for x in range(5) if x != j]:
            bag_train(df, wb[SETS[i]], word_set)

        bag_filter(df)
        bag_composition(df)

        # extracting good and bad words
        goods = []
        bads = []
        for i in range(df.shape[0]):
            if df[-1][i] > 0.5:
                bads.append(df.index[i])
            elif df[1][i] > 0.5:
                goods.append(df.index[i])
        
        print "Set " + str(j + 1) + ": " \
            + str(bag_test(goods, bads, wb[SETS[j]]))

